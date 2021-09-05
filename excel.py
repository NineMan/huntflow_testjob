import logging
import os

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from api import HuntflowAPI

# Состояние кандидата в таблице excel
UPLOADED = 'uploaded'       # загружен в бд
PROCESSED = 'processed'     # полностью обработан

logging.basicConfig(filename="huntflow.log", level=logging.INFO)


class NoNameApplicantError(Exception):
    pass


class NotUploadedCVError(Exception):
    pass


class Applicant:
    """Кандидат на вакансию."""

    def __init__(self, data, db_path):

        self.id = None
        self.fullname = self.get_param(data, 'ФИО')
        self.vacancy = self.get_param(data, 'Должность')
        self.comment = self.get_param(data, 'Комментарий')
        self.status = self.get_param(data, 'Статус')
        self.state = self.get_param(data, 'Состояние')
        self.money = self.get_param(data, 'Ожидания по ЗП')

        self.last_name = self._get_name(0)
        self.first_name = self._get_name(1)
        self.middle_name = None
        self.phones = None
        self.email = None
        self.position = None
        self.company = None
        self.birthday_day = None
        self.birthday_month = None
        self.birthday_year = None
        self.photo_id = None

        # externals
        self.auth_type = 'Native'
        self.account_source = None
        self.body = None
        self.files_id = list()

        self.externals = dict()
        self.set_externals()

        # filename of CV
        self.file_names = self.get_filename_cv(db_path)

    @staticmethod
    def get_param(data, param):
        value = data.get(param)
        return value.strip() if isinstance(value, str) else value

    def _get_name(self, index):
        try:
            return self.fullname.split()[index]
        except Exception:
            return None

    def set_externals(self):
        self.externals = dict(
            data=dict(body=self.body),
            auth_type=self.auth_type,
            account_source=self.account_source,
            files=self.files_id
        )

    def add_files_id(self, file_id):
        dict_id = dict(id=file_id)
        self.files_id.append(dict_id)
        self.externals['files'] = self.files_id

    def get_filename_cv(self, path):
        """Возвращает список имён файлов с резюме кандидата."""

        filenames = []
        for root, dirs, files in os.walk(path, topdown=False):
            for filename in files:
                # Приведение имен файлов с резюме к одному виду с именами в таблице
                clear_filename = clear_text(filename.split('.')[0])
                clear_fullname = clear_text(self.fullname)
                if clear_filename.find(clear_fullname) != -1:
                    filenames.append(f'{root}/{filename}')
        return filenames

    def has_cv(self):
        return self.file_names is not None

    def is_processed(self):
        return self.state == PROCESSED

    def upload_from_cv(self, data: dict):
        """Загрузка данных из распознанного резюме"""

        # self.files_id = data.get('id')
        self.add_files_id(data.get('id'))
        self.body = data.get('text')
        self.set_externals()

        self.auth_type = 'Native'
        self.photo_id = data.get('photo').get('id')

        fields = data.get('fields')
        if fields:
            self.phones = fields.get('phones')
            self.email = fields.get('email')

            name = fields.get('name')
            birthdate = fields.get('birthdate')
            last_experience = fields.get('experience')[0]
            if name:
                self.last_name = name.get('last')
                self.first_name = name.get('first')
                self.middle_name = name.get('middle')
            if birthdate:
                self.birthday_day = birthdate.get('day')
                self.birthday_month = birthdate.get('month')
                self.birthday_year = birthdate.get('year')
            if last_experience:
                self.position = last_experience.get('position')
                self.company = last_experience.get('company')

    def serialize_to_applicant(self) -> dict:

        if not self.last_name or not self.first_name:
            raise NoNameApplicantError('У кандидата должно быть имя и фамилия')

        return dict(
            last_name=self.last_name,
            first_name=self.first_name,
            middle_name=self.middle_name,
            phone=", ".join(self.phones) if self.phones else None,
            email=self.email,
            position=self.position,
            company=self.company,
            money=self.money,
            birthday_day=self.birthday_day,
            birthday_month=self.birthday_month,
            birthday_year=self.birthday_year,
            photo=self.photo_id,
            externals=[
                {
                    "auth_type": "NATIVE",
                    "account_source": None,
                    "data": {
                        "body": self.body
                    },
                    "files": [
                        {
                            "id": self.files_id
                        }
                    ],
                }
            ]
        )

    def print_applicant(self):
        logging.debug(f'====================================================')
        for attr in self.__dict__:
            if attr == 'body':
                continue
            logging.debug(f'{attr} =  {getattr(self, attr)}')

    def __str__(self):
        return self.fullname


class Handler:

    MAX_RETRY_UPLOADED_CV = 2
    MAX_RETRY_UPLOADED_APPLICANT = 2
    MAX_RETRY_UPLOADED_TO_VACANCY = 2

    def __init__(self, config):
        self.cfg = config
        self.api = HuntflowAPI(self.cfg)

    def start(self):
        """Начало обработки кандидата"""

        # Получаем данные из excel
        filename = self.get_excel_filename()
        data = pd.read_excel(filename, sheet_name=self.cfg.EXCEL_SHEET)
        data = data.replace({np.nan: None})

        # Проходим по всем строкам в таблице
        for row in data.iterrows():
            applicant = Applicant(row[1].to_dict(), self.cfg.file_path)
            # Запускаем обработку в если кандидат ещё не обработан
            if applicant.is_processed():
                continue
            self.process_applicant(applicant)
            self.mark_as_processed(row[0])
        print(f'Обработка закончена.')

    def get_excel_filename(self):
        """Возвращает excel файл со списком кандидатов."""

        return self.cfg.file_path / self.cfg.EXCEL_FILENAME

    def process_applicant(self, applicant):
        """Обработка кандидата."""

        logging.info(f'Запускается обработка кандидата {applicant}.')
        print(f'Запускается обработка кандидата {applicant}.')
        # Если есть резюме - пытаемся загрузить его на сайт
        if applicant.has_cv():
            self.upload_cv(applicant)

        # Загружаем данные в базу кандидатов
        self.upload_applicant(applicant)

        # Прицепляем кандидата к вакансии
        self.upload_applicant_to_vacancy(applicant)

        logging.info(f'Кандидат {applicant} обработан.')
        logging.info('')
        print(f'Кандидат {applicant} обработан.')
        print()

    def upload_cv(self, applicant):
        """Загружает CV на сайт, результат загружает в экземпляр кандидата"""

        attempt = 0
        while attempt != self.MAX_RETRY_UPLOADED_CV:
            try:
                logging.info(f'Старт загрузки файла {applicant.file_names[0]}')
                result = self.api.post_file(applicant.file_names[0])
                result.raise_for_status()
                logging.info(f'Резюме загружено в базу. Результат = {result.status_code}')
                cv_data = result.json()
                applicant.upload_from_cv(cv_data)
                return True
            except Exception as error:
                attempt += 1
                logging.error(f'Error uploaded cv, {attempt = }, {error = }')
        return False

    def upload_applicant(self, applicant):
        attempt = 0
        while attempt != self.MAX_RETRY_UPLOADED_APPLICANT:
            try:
                result = self.api.post_applicant(applicant.serialize_to_applicant())
                result.raise_for_status()
                applicant.id = result.json().get('id')
                logging.info(f'Кандидат загружен в базу. Результат = {result.status_code}')
                return True
            except Exception as error:
                attempt += 1
                logging.error(f'Error uploaded applicant, {attempt = }, {error = }')
        return False

    def upload_applicant_to_vacancy(self, applicant):
        attempt = 0
        while attempt != self.MAX_RETRY_UPLOADED_TO_VACANCY:
            try:
                vacancy_id = self.api.list_vacancies.get(applicant.vacancy)
                status_id = self.api.list_statuses.get(applicant.status)
                result = self.api.post_applicant_vacancy(applicant, vacancy_id, status_id)
                logging.info(f'Кандидат прикреплён к вакансии. Результат = {result.status_code}')
                result.raise_for_status()
                return True
            except Exception as error:
                attempt += 1
                logging.error(f'Error uploaded vacancy, {attempt = }, {error = }')
        return False

    def mark_as_processed(self, row):
        """Ставит отметку в дополнительном столбце таблице о выполненной обработке кандидата"""

        filename = self.get_excel_filename()
        workbook = load_workbook(filename)
        worksheet = workbook[self.cfg.EXCEL_SHEET]
        worksheet.cell(row=1, column=6, value='Состояние')
        worksheet.cell(row=row + 2, column=6, value='processed')
        workbook.save(filename)


def clear_text(string):
    """
    Приведение строк с кириллицей к одному виду.

    В кириллице буква "й" может быть представлена двумя вариантами:
     - двумя символами соответствующими символам Юникода '1080' и '774'
     - символом соответствующим символу Юникода '1081'
    Метод приводит строки с кириллицой ко второму варианту.
    """

    string = string.strip()
    old_chars = chr(1080) + chr(774)
    new_chars = chr(1081)
    return string.replace(old_chars, new_chars)


def start_service(cfg):
    Handler(cfg).start()
